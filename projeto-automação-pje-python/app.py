from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

numero_oab = 133864

#abrir no chrome
driver = webdriver.Chrome()#abre o chrome
driver.get("https://pje-consulta-publica.tjmg.jus.br/")#abre a página requerida
sleep(30)#espera por 10 segundos

#digitar numero oab
campo_oab = driver.find_element(By.XPATH,"//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

#selecionar estado
dropdown_estados = driver.find_element(By.XPATH,"//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
opcoes_estados.select_by_visible_text('SP')

#clickar em pesquisar
botao_pesquisar = driver.find_element(By.XPATH,"//input[@id='fPP:searchProcessos']")
botao_pesquisar.click()
sleep(10)

#entrar em cada processo 
processos = driver.find_elements(By.XPATH,"//b[@class='btn-block']")
for processo in processos:
    processo.click()
    sleep(10)
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1920,1080)
    
    #extrair o nº do processo e data da distribuição
    numero_processo = driver.find_elements(By.XPATH,"//div[@class='col-sm-12']")
    numero_processo = numero_processo[0]
    numero_processo= numero_processo.text
    
    data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
    data_distribuicao = data_distribuicao[1]
    data_distribuicao = data_distribuicao.text
    
    #extrair e guardar todas as ultimas movimentações
    movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class, 'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = []
    for movimentacao in movimentacoes:
        lista_movimentacoes.append(movimentacao.text)
    
    workbook = openpyxl.load_workbook("dados.xlsx")
    
    try:
        #codigo para inserir dados em pagina existente
        #acessar página do processo
        pagina_processo = workbook[numero_processo]
        
        #criar nome das colunas
        pagina_processo["A1"].value = "número processo"
        pagina_processo["B1"].value = "data da distribuição"
        pagina_processo["C1"].value = "movimentações"
        
        #adicionar numero do processo
        pagina_processo["A2"].value = numero_processo
        
        #adicionar data de distribuição
        pagina_processo["B2"].value = data_distribuicao
        
        #adicionar movimentações
        for index, row in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=3, max_col=3)):
            for cell in row:
                cell.value = lista_movimentacoes[index]
        
        workbook.save("dados.xlsx")
        driver.close()
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])
        
    except Exception as error:
        #código para criar uma pagina e inserir informações
        workbook.create_sheet(numero_processo)
        
        #acessar página do processo
        pagina_processo = workbook[numero_processo]
        
        #criar nome das colunas
        pagina_processo["A1"].value = "número processo"
        pagina_processo["B1"].value = "data da distribuição"
        pagina_processo["C1"].value = "movimentações"
        
        #adicionar numero do processo
        pagina_processo["A2"].value = numero_processo
        
        #adicionar data de distribuição
        pagina_processo["B2"].value = data_distribuicao
        
        #adicionar movimentações
        for index, row in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=3, max_col=3)):
            for cell in row:
                cell.value = lista_movimentacoes[index]
                
        workbook.save("dados.xlsx")
        driver.close()
        driver.switch_to.window(driver.window_handles[0])